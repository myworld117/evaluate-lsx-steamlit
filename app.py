"""
Đánh giá LSX - Dashboard
========================
Web app hiển thị và phân tích dữ liệu đánh giá Lệnh Sản Xuất.
Chạy: streamlit run app.py
"""
import streamlit as st
import pandas as pd
import numpy as np
import importlib

# Cho phép style dataframe lớn (55K rows × 28 cols ≈ 1.5M cells)
pd.set_option("styler.render.max_elements", 2_000_000)
from pathlib import Path
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pipeline

# Nạp lại pipeline khi app đang chạy để các thay đổi công thức được áp dụng
# ngay sau khi Streamlit rerun, tránh giữ function cũ trong sys.modules.
run_pipeline = importlib.reload(pipeline).process

# ─── Page Config ───────────────────────────────────────────
st.set_page_config(
    page_title="Đánh giá LSX - Dashboard",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Paths ──────────────────────────────────────────────────
DATA_FILE = "TỔNG HỢP LSX tháng 7 (1-24.7).xlsb"

# ─── Color Constants ────────────────────────────────────────
COLOR_RED_BG = "#FFC7CE"
COLOR_RED_FONT = "#9C0006"
COLOR_YELLOW_BG = "#FFEB9C"
COLOR_YELLOW_FONT = "#9C6500"
COLOR_BLUE_BG = "#BDD7EE"
COLOR_BLUE_FONT = "#1F4E79"
COLOR_ORANGE_BG = "#F4B183"
COLOR_PURPLE_BG = "#D9C2EC"
COLOR_GRAY_BG = "#D9D9D9"

# ─── Shorthand for Vietnamese column names (used throughout UI) ───
C = lambda s: s  # placeholder — actual names below

COL_XH = "XH"
COL_DCSX = "DCSX/GC"
COL_MA_LSX = "Mã số lệnh tạo"
COL_NGAY_KC = "Ngày khởi công thực tế"
COL_NGAY_HT = "Ngày hoàn tất thực tế"
COL_MA_SP = "Mã SP"
COL_TEN_VP = "Tên VP"
COL_TINH_TRANG = "Tình trạng lệnh SX"
COL_SL_DU_TINH = "SL dự tính"
COL_SL_THUC_TE = "SL thực tế"
COL_MA_NVL = "Mã NVL"
COL_TEN_VP2 = "Tên VP2"
COL_LUONG_DUNG_TC = "Lượng dùng tiêu chuẩn"
COL_SL_DUNG_THUC = "SL dùng thực"
COL_CHENH_LECH = "Chênh lệch lượng dùng"
COL_PHAN_LOAI = "Phân loại"
COL_NVL_THAY_THE = "NVL thay thế"
COL_CT1 = "Công thức 1 (%)"
COL_CT2 = "Công thức 2 (Tổng chênh lệch)"
COL_CT3 = "Công thức 3 (%)"
COL_CT4 = "Công thức 4 (Tổng chênh lệch)"
COL_CT5 = "Công thức 5 (%)"
COL_CT6 = "Công thức 6 (Tổng chênh lệch)"
COL_CT7 = "Công thức 7 (%)"
COL_CT8 = "Công thức 8 (Tổng chênh lệch)"
COL_LSX_KLL = "LSX không lĩnh liệu"
COL_5402 = "5402 - Phiếu Lĩnh Liệu"
COL_GHI_CHU = "GHI CHÚ"
COL_XH1 = "XH=1"
COL_XH2 = "XH=2"
COL_XH3 = "XH=3"
COL_XHGT3 = "XH>3"
COL_DA_SP = "ĐÁNH GIÁ TRÊN SP"
COL_DA_LSX = "ĐÁNH GIÁ TRÊN LSX"


# ═══════════════════════════════════════════════════════════════
# DATA LOADING (cached)
# ═══════════════════════════════════════════════════════════════

def load_all_data():
    """Load tất cả các sheet dữ liệu nguồn."""
    xlsb = pd.ExcelFile(DATA_FILE, engine="pyxlsb")

    df_mocr = pd.read_excel(xlsb, sheet_name="MOCR27", engine="pyxlsb")
    artifact_cols = [c for c in df_mocr.columns if "Unnamed" in str(c)]
    df_mocr = df_mocr.drop(columns=artifact_cols, errors="ignore")

    df_bom = pd.read_excel(xlsb, sheet_name="BOM ĐẾN 24.07", engine="pyxlsb")
    df_bomr20 = pd.read_excel(xlsb, sheet_name="BOMR20 - NVL TT", engine="pyxlsb")
    df_invr17 = pd.read_excel(xlsb, sheet_name="INVR17", engine="pyxlsb")

    return {
        "mocr": df_mocr,
        "bom": df_bom,
        "bomr20": df_bomr20,
        "invr17": df_invr17,
        "metadata": {
            "n_mocr": len(df_mocr),
            "n_bom": len(df_bom),
            "n_bomr20": len(df_bomr20),
            "n_invr17": len(df_invr17),
        },
    }


# ═══════════════════════════════════════════════════════════════
# UI HELPERS
# ═══════════════════════════════════════════════════════════════

def style_dataframe(df_display: pd.DataFrame):
    """Áp dụng highlight màu cho dataframe."""
    styled = df_display.style

    # Highlight Chênh lệch
    def hl_chenh(val):
        if pd.isna(val) or val == 0: return ""
        if val < 0: return f"background-color: {COLOR_RED_BG}; color: {COLOR_RED_FONT}; font-weight: bold"
        return f"background-color: {COLOR_YELLOW_BG}; color: {COLOR_YELLOW_FONT}"
    styled = styled.map(hl_chenh, subset=[COL_CHENH_LECH])

    # Highlight NVL thay thế
    def hl_nvltt(val):
        if val and str(val).strip(): return f"background-color: {COLOR_BLUE_BG}; color: {COLOR_BLUE_FONT}; font-weight: bold"
        return ""
    styled = styled.map(hl_nvltt, subset=[COL_NVL_THAY_THE])

    # Highlight LSX không lĩnh liệu
    def hl_kll(val):
        if val and "không lĩnh liệu" in str(val).lower():
            return f"background-color: {COLOR_ORANGE_BG}; font-weight: bold"
        return ""
    styled = styled.map(hl_kll, subset=[COL_LSX_KLL])

    # Highlight 5402
    def hl_5402(val):
        if val and str(val).strip(): return f"background-color: {COLOR_PURPLE_BG}; font-weight: bold"
        return ""
    styled = styled.map(hl_5402, subset=[COL_5402])

    # Highlight Ghi chú
    def hl_gc(val):
        if val and str(val).strip(): return f"background-color: {COLOR_GRAY_BG}; font-style: italic"
        return ""
    styled = styled.map(hl_gc, subset=[COL_GHI_CHU])

    # Highlight cột đánh giá
    def hl_danhgia(val):
        if pd.isna(val) or str(val).strip() == "": return ""
        s = str(val).strip()
        if "không đạt" in s.lower(): return f"background-color: {COLOR_RED_BG}; color: {COLOR_RED_FONT}; font-weight: bold"
        if "không lĩnh liệu" in s.lower(): return f"background-color: {COLOR_ORANGE_BG}; font-weight: bold"
        if s == "x": return f"background-color: {COLOR_GRAY_BG}; font-style: italic"
        if "đạt" in s.lower(): return f"background-color: #C6EFCE; color: #006100"
        return ""
    for col in [COL_XH1, COL_XH2, COL_XH3, COL_XHGT3, COL_DA_SP, COL_DA_LSX]:
        styled = styled.map(hl_danhgia, subset=[col])

    # Format số
    styled = styled.format({
        COL_CHENH_LECH: "{:,.3f}",
        COL_LUONG_DUNG_TC: "{:,.3f}",
        COL_SL_DUNG_THUC: "{:,.3f}",
        COL_CT1: "{:.2%}", COL_CT2: "{:,.2f}",
        COL_CT3: "{:.2%}", COL_CT4: "{:,.2f}",
        COL_CT5: "{:.2%}", COL_CT6: "{:,.2f}",
        COL_CT7: "{:.2%}", COL_CT8: "{:,.2f}",
    }, na_rep="-")

    return styled


def export_to_excel(df: pd.DataFrame) -> BytesIO:
    """Xuất DataFrame ra file Excel có highlight màu."""
    wb = Workbook(); ws = wb.active
    ws.title = "LSX Evaluation"

    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_font = Font(color="9C6500")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    orange_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    purple_fill = PatternFill(start_color="D9C2EC", end_color="D9C2EC", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Headers
    for ci, cn in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=cn)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin

    # Column indices
    ci_chenh = df.columns.get_loc(COL_CHENH_LECH) + 1
    ci_nvltt = df.columns.get_loc(COL_NVL_THAY_THE) + 1
    ci_kll = df.columns.get_loc(COL_LSX_KLL) + 1
    ci_5402 = df.columns.get_loc(COL_5402) + 1
    ci_gc = df.columns.get_loc(COL_GHI_CHU) + 1

    for ri, (_, row) in enumerate(df.iterrows()):
        er = ri + 2
        for ci, cn in enumerate(df.columns):
            val = row[cn]
            if pd.isna(val): val = ""
            ws.cell(row=er, column=ci + 1, value=val).border = thin

        cv = row[COL_CHENH_LECH]
        if pd.notna(cv) and cv != 0:
            c = ws.cell(row=er, column=ci_chenh)
            if cv < 0: c.fill = red_fill; c.font = red_font
            else: c.fill = yellow_fill; c.font = yellow_font

        if row[COL_NVL_THAY_THE] and str(row[COL_NVL_THAY_THE]).strip():
            ws.cell(row=er, column=ci_nvltt).fill = blue_fill
        if row[COL_LSX_KLL] and "không lĩnh liệu" in str(row[COL_LSX_KLL]).lower():
            ws.cell(row=er, column=ci_kll).fill = orange_fill
        if row[COL_5402] and str(row[COL_5402]).strip():
            ws.cell(row=er, column=ci_5402).fill = purple_fill
        if row[COL_GHI_CHU] and str(row[COL_GHI_CHU]).strip():
            ws.cell(row=er, column=ci_gc).fill = gray_fill

    # Column widths
    for ci, cn in enumerate(df.columns):
        ml = max(len(str(cn)), df[cn].astype(str).str.len().max() if len(df) > 0 else 0)
        ws.column_dimensions[ws.cell(row=1, column=ci + 1).column_letter].width = min(ml + 4, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(df.columns)).column_letter}{len(df) + 1}"

    out = BytesIO(); wb.save(out); out.seek(0)
    return out


# ═══════════════════════════════════════════════════════════════
# MAIN UI
# ═══════════════════════════════════════════════════════════════

def main():
    # Xóa cache cũ để đảm bảo chạy code mới nhất
    st.cache_data.clear()

    st.title("🔍 Đánh giá LSX — Dashboard")
    st.caption(
        f"File dữ liệu: `{DATA_FILE}` — Sheet `MOCR27` — "
        f"Cập nhật: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )

    # ── Load ────────────────────────────────────────────────
    with st.spinner("🔄 Đang tải dữ liệu từ các sheet..."):
        try:
            data = load_all_data()
        except Exception as e:
            st.error(f"❌ Lỗi tải dữ liệu: {e}")
            st.info(f"Đảm bảo file `{DATA_FILE}` có trong thư mục hiện tại.")
            return

    meta = data["metadata"]

    # ── Run pipeline ────────────────────────────────────────
    with st.spinner("⚙️ Đang xử lý pipeline (join BOM, tính chênh lệch, công thức...)"):
        try:
            result = run_pipeline(data)
            n_sub = int((result[COL_NVL_THAY_THE].str.strip() != "").sum())
        except Exception as e:
            st.error(f"❌ Lỗi xử lý pipeline: {e}")
            import traceback; st.code(traceback.format_exc())
            return

    st.success(
        f"✅ Đã tải: MOCR27 ({meta['n_mocr']:,} dòng) | "
        f"BOM ({meta['n_bom']:,} dòng) | "
        f"BOMR20 ({meta['n_bomr20']:,} dòng) | "
        f"INVR17 ({meta['n_invr17']:,} dòng) | "
        f"🔵 NVL thay thế: **{n_sub:,}**"
    )

    # ── Sidebar ─────────────────────────────────────────────
    with st.sidebar:
        st.header("🎯 Bộ lọc")
        st.metric("Tổng số dòng", f"{len(result):,}")
        st.metric("LSX duy nhất", f"{result[COL_MA_LSX].nunique():,}")
        st.metric("NVL duy nhất", f"{result[COL_MA_NVL].nunique():,}")
        st.divider()

        tt_opts = ["Tất cả"] + sorted(result[COL_TINH_TRANG].dropna().unique().tolist())
        sel_tt = st.selectbox("Tình trạng lệnh SX", tt_opts)

        xh_opts = ["Tất cả"] + sorted(result[COL_XH].dropna().unique().tolist(), key=lambda x: int(x) if x.isdigit() else 999)
        sel_xh = st.selectbox("Số lần xuất hiện (XH)", xh_opts)

        pl_opts = ["Tất cả"] + sorted(result[COL_PHAN_LOAI].dropna().unique().tolist())
        sel_pl = st.selectbox("Phân loại NVL", pl_opts)

        st.subheader("Khoảng chênh lệch")
        mn, mx = float(result[COL_CHENH_LECH].min()), float(result[COL_CHENH_LECH].max())
        chenh_range = st.slider("Chênh lệch", mn, mx, (mn, mx), 0.01, format="%.2f")

        search = st.text_input("🔎 Tìm kiếm", placeholder="Mã LSX, Mã SP, Mã NVL, Tên VP...")

        show_sub = st.checkbox("Chỉ hiện NVL thay thế (*)")
        show_kll = st.checkbox("Chỉ hiện LSX không lĩnh liệu")
        show_5402 = st.checkbox("Chỉ hiện phiếu lĩnh vượt 5402")
        show_gc = st.checkbox("Chỉ hiện dòng có ghi chú")

        st.divider()
        st.subheader("📥 Export")

        if st.button("⬇️ Tải Excel (có highlight)", width="stretch"):
            xl = export_to_excel(result)
            st.download_button("📥 Click để tải Excel", xl,
                file_name=f"LSX_Evaluation_{datetime.now():%Y%m%d_%H%M}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch")

        csv = result.to_csv(index=False).encode("utf-8-sig")
        st.download_button("⬇️ Tải CSV", csv,
            file_name=f"LSX_Evaluation_{datetime.now():%Y%m%d_%H%M}.csv",
            mime="text/csv", width="stretch")

    # ── Apply filters ───────────────────────────────────────
    f = result.copy()
    if sel_tt != "Tất cả": f = f[f[COL_TINH_TRANG] == sel_tt]
    if sel_xh != "Tất cả": f = f[f[COL_XH] == sel_xh]
    if sel_pl != "Tất cả": f = f[f[COL_PHAN_LOAI] == sel_pl]
    f = f[(f[COL_CHENH_LECH] >= chenh_range[0]) & (f[COL_CHENH_LECH] <= chenh_range[1])]

    if search:
        txt = search.lower()
        m = (f[COL_MA_LSX].fillna("").astype(str).str.lower().str.contains(txt, na=False) |
             f[COL_MA_SP].fillna("").astype(str).str.lower().str.contains(txt, na=False) |
             f[COL_MA_NVL].fillna("").astype(str).str.lower().str.contains(txt, na=False) |
             f[COL_TEN_VP].fillna("").astype(str).str.lower().str.contains(txt, na=False) |
             f[COL_TEN_VP2].fillna("").astype(str).str.lower().str.contains(txt, na=False))
        f = f[m]

    if show_sub: f = f[f[COL_NVL_THAY_THE].str.strip() != ""]
    if show_kll: f = f[f[COL_LSX_KLL].str.contains("không lĩnh liệu", na=False)]
    if show_5402: f = f[f[COL_5402].str.strip() != ""]
    if show_gc: f = f[f[COL_GHI_CHU].str.strip() != ""]

    # ── Tabs ────────────────────────────────────────────────
    t1, t2, t3 = st.tabs(["📊 Tổng quan", "📋 Dữ liệu chi tiết", "⚠️ Bất thường"])

    # ── TAB 1: Tổng quan ────────────────────────────────────
    with t1:
        st.header("Tổng quan dữ liệu")
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("🔴 Vượt định mức", f"{(f[COL_CHENH_LECH] < 0).sum():,}")
        c2.metric("🟢 Đúng định mức", f"{(f[COL_CHENH_LECH] == 0).sum():,}")
        c3.metric("🟡 Dưới định mức", f"{(f[COL_CHENH_LECH] > 0).sum():,}")
        c4.metric("🔵 NVL thay thế", f"{(f[COL_NVL_THAY_THE].str.strip() != '').sum():,}")
        c5.metric("🟠 Không lĩnh liệu", f"{f[COL_LSX_KLL].str.contains('không lĩnh', na=False).sum():,}")
        st.divider()

        cl, cr = st.columns(2)
        with cl:
            st.subheader("Phân bố chênh lệch")
            fig = px.histogram(f, x=COL_CHENH_LECH, nbins=100,
                color_discrete_sequence=["#4472C4"],
                labels={COL_CHENH_LECH: "Chênh lệch", "count": "Số dòng"})
            fig.add_vline(x=0, line_dash="dash", line_color="red", annotation_text="Đúng định mức")
            fig.update_layout(height=400)
            st.plotly_chart(fig, width="stretch")
        with cr:
            st.subheader("Phân loại NVL")
            pc = f[COL_PHAN_LOAI].value_counts().reset_index()
            pc.columns = ["Loại", "SL"]
            fig = px.pie(pc, values="SL", names="Loại", hole=0.4)
            fig.update_layout(height=400)
            st.plotly_chart(fig, width="stretch")

        cl2, cr2 = st.columns(2)
        with cl2:
            st.subheader("Chênh lệch theo phân loại NVL")
            fig = px.box(f, x=COL_PHAN_LOAI, y=COL_CHENH_LECH, color=COL_PHAN_LOAI,
                labels={COL_CHENH_LECH: "Chênh lệch", COL_PHAN_LOAI: "Phân loại"})
            fig.update_layout(height=400, showlegend=False, xaxis_tickangle=-45)
            st.plotly_chart(fig, width="stretch")
        with cr2:
            st.subheader("Top 10 NVL chênh lệch lớn nhất (|value|)")
            t10 = f.copy()
            t10["abs"] = t10[COL_CHENH_LECH].abs()
            t10 = t10.nlargest(10, "abs")[[COL_MA_NVL, COL_TEN_VP2, COL_CHENH_LECH, COL_PHAN_LOAI]]
            t10 = t10.drop_duplicates(subset=[COL_MA_NVL])
            fig = px.bar(t10, x=COL_CHENH_LECH, y=COL_MA_NVL, orientation="h",
                color=COL_CHENH_LECH, color_continuous_scale=["red", "lightgray", "orange"],
                labels={COL_CHENH_LECH: "Chênh lệch", COL_MA_NVL: "Mã NVL"},
                hover_data=[COL_TEN_VP2, COL_PHAN_LOAI])
            fig.update_layout(height=400, yaxis={"categoryorder": "total ascending"})
            st.plotly_chart(fig, width="stretch")

    # ── TAB 2: Dữ liệu chi tiết ─────────────────────────────
    with t2:
        st.header(f"Dữ liệu chi tiết ({len(f):,} dòng)")
        st.caption("🟡 Vàng = dưới định mức | 🔴 Đỏ = vượt định mức | 🔵 Xanh = NVL thay thế | 🟠 Cam = Không lĩnh liệu | 🟣 Tím = 5402")

        disp_cols = [
            COL_XH, COL_DCSX, COL_MA_LSX, COL_NGAY_KC, COL_NGAY_HT,
            COL_MA_SP, COL_TEN_VP, COL_TINH_TRANG, COL_SL_DU_TINH, COL_SL_THUC_TE,
            COL_MA_NVL, COL_TEN_VP2, COL_LUONG_DUNG_TC, COL_SL_DUNG_THUC, COL_CHENH_LECH,
            COL_PHAN_LOAI, COL_NVL_THAY_THE,
            COL_CT2, COL_CT1, COL_CT4, COL_CT3,
            COL_CT6, COL_CT5, COL_CT8, COL_CT7,
            COL_LSX_KLL,
            COL_XH1, COL_XH2, COL_XH3, COL_XHGT3,
            COL_DA_SP, COL_5402, COL_DA_LSX,
            COL_GHI_CHU,
        ]
        df_disp = f[disp_cols]
        styled_df = style_dataframe(df_disp)

        st.dataframe(styled_df, width="stretch", height=700, hide_index=True,
            column_config={
                COL_CHENH_LECH: st.column_config.NumberColumn(format="%.3f"),
                COL_LUONG_DUNG_TC: st.column_config.NumberColumn(format="%.3f"),
                COL_CT1: st.column_config.NumberColumn(format="percent"),
                COL_CT2: st.column_config.NumberColumn(format="%.2f"),
                COL_CT3: st.column_config.NumberColumn(format="percent"),
                COL_CT4: st.column_config.NumberColumn(format="%.2f"),
                COL_CT5: st.column_config.NumberColumn(format="percent"),
                COL_CT6: st.column_config.NumberColumn(format="%.2f"),
                COL_CT7: st.column_config.NumberColumn(format="percent"),
                COL_CT8: st.column_config.NumberColumn(format="%.2f"),
            })
        st.caption(f"Hiển thị {len(f):,} / {len(result):,} dòng")

    # ── TAB 3: Bất thường ───────────────────────────────────
    with t3:
        st.header("⚠️ Dữ liệu bất thường cần chú ý")
        am = ((f[COL_CHENH_LECH] < 0) |
              (f[COL_NVL_THAY_THE].str.strip() != "") |
              (f[COL_LSX_KLL].str.contains("không lĩnh", na=False)) |
              (f[COL_5402].str.strip() != "") |
              (f[COL_GHI_CHU].str.strip() != ""))
        anom = f[am].copy()
        st.metric("Tổng số dòng bất thường", f"{len(anom):,}")

        if len(anom) > 0:
            reasons = []
            for _, row in anom.iterrows():
                r = []
                if row[COL_CHENH_LECH] < 0: r.append("🔴 Vượt định mức")
                if row[COL_NVL_THAY_THE] and str(row[COL_NVL_THAY_THE]).strip(): r.append("🔵 NVL thay thế")
                if "không lĩnh" in str(row[COL_LSX_KLL]).lower(): r.append("🟠 Không lĩnh liệu")
                if str(row[COL_5402]).strip(): r.append("🟣 Phiếu 5402")
                if str(row[COL_GHI_CHU]).strip(): r.append("⚫ Ghi chú")
                reasons.append(", ".join(r))
            anom["Loại bất thường"] = reasons

            ad = anom[[COL_XH, COL_MA_LSX, COL_MA_SP, COL_MA_NVL, COL_TEN_VP2,
                        COL_CHENH_LECH, COL_PHAN_LOAI, "Loại bất thường",
                        COL_LSX_KLL, COL_5402, COL_GHI_CHU]]
            st.dataframe(ad, width="stretch", height=600, hide_index=True,
                column_config={COL_CHENH_LECH: st.column_config.NumberColumn(format="%.3f")})

            st.subheader("Phân bố loại bất thường")
            rc = pd.Series(reasons).value_counts().reset_index()
            rc.columns = ["Loại", "SL"]
            fig = px.bar(rc, x="SL", y="Loại", orientation="h", color="SL")
            fig.update_layout(height=300)
            st.plotly_chart(fig, width="stretch")
        else:
            st.info("Không có dòng bất thường nào.")

    # ── Footer ──────────────────────────────────────────────
    st.divider()
    st.caption(
        f"🤖 Pipeline: MOCR27 → Join BOM + BOMR20 → Tính chênh lệch → "
        f"Phân loại → 8 CT → 5402 check | "
        f"Tổng: {len(result):,} dòng | Filtered: {len(f):,} dòng"
    )


if __name__ == "__main__":
    main()
